"""Tests for the NFR Excel workbook generator.

Validates REPORT-03: correct sheet structure, cell values from fixture data,
sheet name truncation, and formula injection mitigation (T-136-09).
"""

from __future__ import annotations

import json
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest

from aila.modules.sbd_nfr.reporting.excel_service import _safe_cell_value

# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

_FIXTURES_DIR = Path(__file__).resolve().parent / "fixtures"


@pytest.fixture
def golden_resolution() -> dict:
    """Load golden_resolution_response.json fixture."""
    return json.loads((_FIXTURES_DIR / "golden_resolution_response.json").read_text())


@pytest.fixture
def golden_answers() -> list[dict]:
    """Load golden_session_answers.json fixture."""
    return json.loads((_FIXTURES_DIR / "golden_session_answers.json").read_text())


# ---------------------------------------------------------------------------
# _safe_cell_value: formula injection mitigation (T-136-09)
# ---------------------------------------------------------------------------


def test_safe_cell_value_prefix_equals():
    """Values starting with '=' must be prefixed with single-quote."""
    assert _safe_cell_value("=SUM(A1:A10)") == "'=SUM(A1:A10)"


def test_safe_cell_value_prefix_plus():
    """Values starting with '+' must be prefixed with single-quote."""
    assert _safe_cell_value("+447700000000") == "'+447700000000"


def test_safe_cell_value_prefix_minus():
    """Values starting with '-' must be prefixed with single-quote."""
    assert _safe_cell_value("-1 days") == "'-1 days"


def test_safe_cell_value_prefix_at():
    """Values starting with '@' must be prefixed with single-quote."""
    assert _safe_cell_value("@admin") == "'@admin"


def test_safe_cell_value_plain_text_unchanged():
    """Plain text values that do not start with formula chars are returned unchanged."""
    assert _safe_cell_value("External (APN; customer facing)") == "External (APN; customer facing)"
    assert _safe_cell_value("Yes") == "Yes"
    assert _safe_cell_value("") == ""


def test_safe_cell_value_none_safe():
    """Empty string input returns empty string without error."""
    assert _safe_cell_value("") == ""


def test_safe_cell_value_preserves_content_after_prefix():
    """The original content is preserved after the single-quote prefix."""
    original = "=MALICIOUS()"
    result = _safe_cell_value(original)
    assert result.startswith("'")
    assert result[1:] == original


# ---------------------------------------------------------------------------
# Workbook builder: direct unit tests via _build_workbook
# ---------------------------------------------------------------------------


def _make_mock_session():
    """Create a minimal mock session object for _build_workbook."""

    class _MockSession:
        id = "test-session-001"
        project_name = "Test NFR Project"
        status = "resolved"
        requestor_name = "Jane Doe"
        requestor_email = "jane@example.com"
        business_unit = "Engineering"

    return _MockSession()


def _make_mock_results(golden_resolution: dict) -> list:
    """Create mock result objects from golden fixture data."""
    from datetime import datetime, timezone

    class _MockResult:
        def __init__(self, comp: dict):
            self.subtask_key = comp["subtask_key"]
            self.classification = comp["classification"]
            self.confidence = comp["confidence"]
            self.reasoning = comp.get("reasoning", "")
            self.cited_question_ids_json = json.dumps(comp.get("cited_question_ids", []))
            self.resolved_at = datetime.now(timezone.utc)

    return [_MockResult(c) for c in golden_resolution["components"]]


def _make_mock_answers(golden_answers: list[dict]) -> list:
    """Create mock answer objects from golden_session_answers fixture."""

    class _MockAnswer:
        def __init__(self, ans: dict):
            self.question_id = ans["question_id"]
            self.answer_value = ans["answer_value"]
            self.note_text = ans.get("note_text")

    return [_MockAnswer(a) for a in golden_answers]


def _make_mock_sections() -> list:
    """Create a minimal set of mock section objects."""

    class _MockSection:
        def __init__(self, key: str, label: str, order: int):
            self.section_key = key
            self.label = label
            self.display_order = order
            self.is_active = True

    return [
        _MockSection("base_questionnaire", "Base Questionnaire", 1),
        _MockSection("hygiene_essentials", "Hygiene Essentials", 2),
        _MockSection("data_protection", "Data Protection", 3),
        _MockSection("logging_monitoring", "Logging Monitoring", 4),
        _MockSection("user_accounts", "User Accounts", 5),
        _MockSection("apis", "APIs", 6),
        _MockSection("supplier_3rd_party", "Supplier 3rd Party", 7),
        _MockSection("web_mobile", "Web Mobile", 8),
    ]


def _make_mock_questions() -> list:
    """Create minimal mock question objects for common question IDs."""

    class _MockQuestion:
        def __init__(self, qid: str, label: str, subgroup_id: str = "sg-1"):
            self.id = qid
            self.label = label
            self.subgroup_id = subgroup_id
            self.is_active = True

    return [
        _MockQuestion("SCOPE-01", "What is the service exposure type?"),
        _MockQuestion("SCOPE-02", "Is this a new production service?"),
        _MockQuestion("SCOPE-03", "What is the deployment platform?"),
        _MockQuestion("SCOPE-04", "Is this a managed cloud service?"),
        _MockQuestion("SCOPE-05", "Does the service process user events?"),
        _MockQuestion("SCOPE-06", "Does the service store data?"),
        _MockQuestion("SCOPE-07", "Are there privileged endpoints?"),
        _MockQuestion("SCOPE-08", "Does the service integrate with third parties?"),
        _MockQuestion("HYGN-01", "Has DMZ placement been validated?"),
        _MockQuestion("HYGN-03", "Does the service require network access points?", "sg-2"),
        _MockQuestion("HYGN-05", "Is container security scanning required?", "sg-2"),
        _MockQuestion("HYGN-06", "Has DAST been performed previously?", "sg-2"),
        _MockQuestion("HYGN-08", "Is SIEM log forwarding required?", "sg-2"),
        _MockQuestion("DATA-01", "Does the service store PII?", "sg-3"),
        _MockQuestion("DATA-03", "Is data shared with third parties?", "sg-3"),
        _MockQuestion("VENDOR-01", "Does the service use third-party vendors?", "sg-7"),
    ]


def test_summary_sheet_exists(golden_resolution, golden_answers):
    """Workbook built from fixture data must contain a 'Summary' sheet (REPORT-03)."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    assert "Summary" in wb.sheetnames


def test_summary_sheet_has_component_rows(golden_resolution, golden_answers):
    """Summary sheet must have at least as many data rows as components in the fixture."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    component_count = len(golden_resolution["components"])
    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    ws = wb["Summary"]
    # Row 1 is the header; subsequent rows are data rows
    data_rows = ws.max_row - 1
    assert data_rows >= component_count, (
        f"Expected at least {component_count} data rows, got {data_rows}"
    )


def test_assessment_info_sheet_exists(golden_resolution, golden_answers):
    """Workbook must include an 'Assessment Info' sheet with session metadata."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    assert "Assessment Info" in wb.sheetnames


def test_workbook_has_at_least_two_sheets(golden_resolution, golden_answers):
    """Workbook must have at least 2 sheets: Summary and Assessment Info."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    assert len(wb.sheetnames) >= 2


def test_workbook_summary_has_header_row(golden_resolution, golden_answers):
    """Summary sheet row 1 must contain the column headers."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    ws = wb["Summary"]
    header_values = [ws.cell(row=1, column=c).value for c in range(1, 6)]
    assert "Component" in header_values
    assert "Classification" in header_values


def test_workbook_section_sheet_has_answer_data(golden_resolution, golden_answers):
    """At least one per-section sheet should exist when answers cover multiple sections."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    # Expect at least one section sheet beyond Summary and Assessment Info
    assert len(wb.sheetnames) >= 3, (
        f"Expected at least 3 sheets, got: {wb.sheetnames}"
    )


# ---------------------------------------------------------------------------
# Sheet name truncation (Excel 31-char limit — Pitfall 2)
# ---------------------------------------------------------------------------


def test_sheet_name_truncation_to_31_chars():
    """Section labels longer than 31 chars must produce a sheet name of at most 31 chars.

    Tests the truncation logic by using a section whose label exceeds 31 characters
    and verifying the resulting workbook has no sheet name over the Excel limit.
    Uses SCOPE prefix (maps to base_questionnaire) so the section is matched.
    """
    from datetime import datetime, timezone

    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    class _LongLabelSection:
        # This key matches SCOPE prefix in _PREFIX_TO_SECTION (mapped to base_questionnaire)
        section_key = "base_questionnaire"
        label = "This Is An Extremely Long Section Label That Exceeds Thirty-One Characters"
        display_order = 1
        is_active = True

    class _MockAnswer:
        question_id = "SCOPE-01"
        answer_value = "Yes"
        note_text = None

    class _MockQuestion:
        id = "SCOPE-01"
        label = "What is the exposure type?"
        subgroup_id = "sg-1"
        is_active = True

    class _MockResult:
        subtask_key = "network_segment_placement"
        classification = "triggered"
        confidence = 0.90
        reasoning = "External service."
        cited_question_ids_json = "[]"
        resolved_at = datetime.now(timezone.utc)

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=[_MockResult()],
        answers=[_MockAnswer()],
        sections=[_LongLabelSection()],
        questions=[_MockQuestion()],
    )
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    # Verify no sheet name exceeds 31 characters (Excel hard limit)
    for sheet_name in wb.sheetnames:
        assert len(sheet_name) <= 31, f"Sheet name '{sheet_name}' exceeds 31 chars"
    # The long label must have been truncated to exactly 31 chars
    assert any(
        name.startswith("This Is An Extremely Long Secti")
        for name in wb.sheetnames
    ), f"Expected truncated sheet name, got: {wb.sheetnames}"


# ---------------------------------------------------------------------------
# Workbook bytes are valid XLSX
# ---------------------------------------------------------------------------


def test_workbook_bytes_are_valid_xlsx(golden_resolution, golden_answers):
    """The bytes returned by _build_workbook must be parseable as a valid XLSX file."""
    from aila.modules.sbd_nfr.reporting.excel_service import _build_workbook

    wb_bytes = _build_workbook(
        session=_make_mock_session(),
        results=_make_mock_results(golden_resolution),
        answers=_make_mock_answers(golden_answers),
        sections=_make_mock_sections(),
        questions=_make_mock_questions(),
    )
    # openpyxl.load_workbook raises if bytes are not a valid XLSX
    wb = openpyxl.load_workbook(BytesIO(wb_bytes))
    assert wb is not None
